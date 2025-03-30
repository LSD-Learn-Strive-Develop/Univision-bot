import re
import openpyxl
import logging

from aiogram import Router, F
from aiogram.types import Message, ReplyKeyboardRemove, ReplyKeyboardMarkup, KeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram_dialog import DialogManager, StartMode
from fluentogram import TranslatorRunner

from bot.config_data.config import db, admins
from bot.states.start import StartSG

other_router = Router()


class Actions(StatesGroup):
    waiting_for_press_start_button = State()
    waiting_for_registration = State()
    waiting_for_votion = State()
    waiting_for_language = State()


@other_router.message(F.document)
async def get_mails(msg: Message, state: FSMContext, i18n: TranslatorRunner):
    if msg.from_user.id not in admins:
        await msg.answer('Вам нельзя загружать документы :)')
        return
    
    document = msg.document
    path = 'correct_mail.xlsx'
    await msg.bot.download(document, path)
    
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for i in range(2, sheet_obj.max_row + 1):
        try:
            faculty = sheet_obj.cell(row = i, column = 1).value
            mail = sheet_obj.cell(row = i, column = 2).value
            mail = mail[:8]
            country = sheet_obj.cell(row = i, column = 4).value

            data = {'faculty': faculty, 'mail': mail}
            if country != 'Россия':
                data['kio'] = True

            db.users.insert_one(data)
        except Exception as e:
            print(e)

    await msg.answer('Загружено')


@other_router.message(Actions.waiting_for_registration)
async def registration_start(msg: Message, state: FSMContext, i18n: TranslatorRunner, dialog_manager: DialogManager):
    if msg.text == i18n.button.employee():
        await state.update_data(status='employee')
        
        await msg.answer(i18n.message.st_format(), reply_markup=ReplyKeyboardRemove())

    else:
        text = msg.text.lower()
        result = re.match(r'st[0-9]{6,6}', text)
        if result != None and len(text) == 8:
            data = {'tg_id': msg.from_user.id}

            user_data = await state.get_data()
            
            if 'status' in user_data:
                #проверить что такой st почты нет
                if db.users.find_one({'mail': text}):
                    db.users.update_one({'mail': text}, {'$set': data})
                else:
                    data['status'] = 'employee'
                    data['mail'] = text
                    db.users.insert_one(data)
            else:
                if db.users.find_one({'mail': text}): 
                    db.users.update_one({'mail': text}, {'$set': data})
                else:
                    await msg.answer(i18n.message.invalid_format()) 

            await dialog_manager.start(state=StartSG.start, mode=StartMode.RESET_STACK)
            await state.clear()
        else:
            await msg.answer(i18n.message.invalid_format())


@other_router.message(Actions.waiting_for_press_start_button)
async def press_start_button(msg: Message, state: FSMContext, i18n: TranslatorRunner, dialog_manager: DialogManager):
    user = db.users.find_one({'tg_id': msg.from_user.id})
    if not user:
        kb = [[KeyboardButton(text=i18n.button.employee())]]
        keyboard = ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True, one_time_keyboard=True)

        await msg.answer(i18n.message.registration(), reply_markup=keyboard)
        await state.set_state(Actions.waiting_for_registration.state)
    else:
        await dialog_manager.start(state=StartSG.start, mode=StartMode.RESET_STACK)
        await state.clear()


@other_router.message()
async def start_logic(msg: Message, state: FSMContext, i18n: TranslatorRunner):
    kb = [[KeyboardButton(text=i18n.button.vote())]]
    keyboard = ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)

    info_db = db.info.find_one()
    if not info_db or info_db['active_event'] == 0:
        await msg.answer(i18n.message.voting_not_running(), reply_markup=keyboard)
        return

    user = db.users.find_one({'tg_id': msg.chat.id})

    if user and 'event_' + str(info_db['active_event']) in user:
        await msg.answer(i18n.message.already_voted(), reply_markup=keyboard)
        return

    await msg.answer(i18n.message.hello(), reply_markup=keyboard)

    await state.set_state(Actions.waiting_for_press_start_button.state)