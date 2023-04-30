import config
import messages
import buttons
import messages_fs
import buttons_fs

import asyncio
import re
import csv
import pandas as pd
import openpyxl

import pymongo
from pymongo import MongoClient

from aiogram import Bot, Dispatcher, executor, types

from aiogram.contrib.fsm_storage.mongo import MongoStorage
from aiogram.utils.executor import start_webhook
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup


WEBHOOK_HOST = 'https://pmpu.site'
WEBHOOK_PATH = config.path
WEBHOOK_URL = f"{WEBHOOK_HOST}{WEBHOOK_PATH}"

WEBAPP_HOST = '127.0.0.1'
WEBAPP_PORT = config.port

# mongo_pass = config.mongo_pass
# mongo_db = config.mongo_db

client = MongoClient('localhost', 27017)
db = client[config.mongo_db_name]

bot = Bot(token=config.token)

OK = '✅'
NOK = '❌'

info = db.info.find_one()['active_event']
admins = [52899166, 248603604, 294062257]

class Actions(StatesGroup):
    waiting_for_press_start_button = State()
    waiting_for_registration = State()
    waiting_for_votion = State()


async def on_startup(dp):
    await bot.set_webhook(WEBHOOK_URL)


async def on_shutdown(dp):
    await bot.delete_webhook()


async def vote_kb(msg):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    row = [await get_button(msg, 'vote')]
    keyboard.add(*row)

    return keyboard


async def status_selection(msg):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    row1 = [await get_button(msg, 'employee')]
    # row2 = [await get_button(msg, 'foreign_student')]
    # keyboard.add(*row1).add(*row2)
    keyboard.add(*row1)

    return keyboard


async def get_message(msg, key):
    status = 'student'
    try:
        user = db.users.find_one({'chat_id': msg.chat.id})
        status = user['status']
    except Exception as e:
        print(e)

    if status != 'foreign_student':
        return messages.d[key]
    else:
        return messages_fs.d[key]


async def get_button(msg, key):
    status = 'student'
    try:
        user = db.users.find_one({'chat_id': msg.chat.id})
        status = user['status']
    except Exception as e:
        print(e)

    if status != 'foreign_student':
        return buttons.d[key]
    else:
        return buttons_fs.d[key]


async def add_user_to_db(msg: types.Message, state: FSMContext):
    users = msg.text.split('\n')
    for user in users:
        chat_id = user.split()[0]
        st_mail = user.split()[1]
        db.users.insert_one({'chat_id': int(chat_id), 'st_main': st_mail})

    await msg.answer('Пользователи добавлены')


async def show(msg: types.Message):
    faculties = db.faculties.find({})
    text = ''
    for faculty in faculties:
        text += faculty['faculty'] + '\n'

    await msg.answer(text)


async def add_faculty(msg: types.Message):
    # db.faculties.insert_one({'name': msg.get_args()})

    rows = msg.get_args().split('\n')
    for row in rows:
        db.faculties.insert_one({'faculty': row.split('$')[0], 
                                 'shot_faculty': row.split('$')[1]})

    await msg.answer('Факультет добавлен')


async def del_faculty(msg: types.Message):
    db.faculties.delete_one({'faculty': msg.get_args()})
    await msg.answer('Факультет удален')


async def event_start(msg: types.Message):
    try:
        if msg.from_user.id not in admins:
            await msg.answer('У вас нет прав на это действие')
            return

        number = int(msg.get_args())
        print(number)

        global info
        info = number

        db.info.update_one(
            {'active_event': 0},
            {'$set': {'active_event': number}}
        )
        await msg.answer('Голосование запущено')
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def event_stop(msg: types.Message):
    try:
        if msg.from_user.id not in admins:
            await msg.answer('У вас нет прав на это действие')
            return

        number = int(msg.get_args())
        print(number)

        global info
        info = 0

        db.info.update_one(
            {'active_event': number},
            {'$set': {'active_event': 0}}
        )
        await msg.answer('Голосование остановлено')
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def get_result(msg: types.Message):
    try:
        if msg.from_user.id not in admins:
            await msg.answer('У вас нет прав на это действие')
            return

        args = str(msg.get_args())
        number_event = args.split()[0]
        flag = False
        if len(args.split()) > 1:
            flag = bool(args.split()[1])
        
        print(flag)
        filename = 'result_' + number_event + '.xlsx'
        
        all_data = []
        for user in db.users.find({ 'event_' + number_event + '.0': {'$exists': True}}):
            faculty = 'Сотрудник'
            if user['faculty'] != '':
                faculty = user['faculty']
            
            if user['status'] == 'foreign_student':
                faculty = 'КИО'

            data = [user['chat_id'], faculty]

            for i in range(3):
                try:
                    data.append(user['event_' + number_event][i])
                except Exception as e:
                    data.append('')
            
            if flag:
                data.append(user['st_main'])
                status = ''
                if 'status' in user:
                    status = user['status']
                data.append(status)

            all_data.append(data)

        columns = ['chat_id', 'От кого', '1', '2', '3']
        if flag:
            columns.append('st_mail')
            columns.append('status')

        df = pd.DataFrame(all_data, columns=columns)
        df.to_excel(filename, index=False, )

        await msg.reply_document(open(filename, 'rb'))
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def add_correct_mails(msg: types.Message):
    path = 'inostr.xlsx'
    # path = 'correct_mail.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for i in range(2, 4198):
        first = sheet_obj.cell(row = i, column = 1).value
        second = sheet_obj.cell(row = i, column = 2).value
        db.foreign.insert_one({'mail': second})
    
    print('Success')
        

async def push_kio(msg: types.Message):
    chat_ids = msg.get_args().split('\n')
    for chat_id in chat_ids:
        t = int(chat_id.split(',')[0])
        db.users.update_one(
            {'chat_id': t},
            {'$set': {'status': 'foreign_student'}}
        )


async def push_kio_from_db(msg: types.Message):
    for obj in db.foreign.find():
        mail = obj['mail']
        if db.users.find_one({'st_main': mail}):
            db.users.update_one(
                {'st_main': mail},
                {'$set': {'status': 'foreign_student'}}
            )

            print(mail)
    
    print('Success')


async def check_user(msg: types.Message):
    try:
        mail = msg.get_args()

        user = db.users.find_one({'st_main': mail})

        if user:
            await msg.answer(str(user))
        else:
            await msg.answer('Такого пользователя нет в базе')
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def del_user(msg: types.Message):
    try:
        mail = msg.get_args()

        db.users.delete_one({'st_main': mail})
        db.users.insert_one({'chat_id': msg.from_user.id, 'st_main': mail})

        await msg.answer('Удалён')
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def del_user_all(msg: types.Message):
    try:
        mail = msg.get_args()

        db.users.delete_one({'st_main': mail})

        await msg.answer('Удалён')
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def press_start_button(msg: types.Message, state: FSMContext):
    try:
        global info
        if info == 0:
            await msg.answer(await get_message(msg, 'voting_end'))
            await state.finish()
            return

        chat_id = msg.from_user.id

        user = db.users.find_one({'chat_id': chat_id})
        if not user:
            db.users.insert_one({'chat_id': chat_id})

            await msg.answer(await get_message(msg, 'registration'), reply_markup=await status_selection(msg))
            await state.set_state(Actions.waiting_for_registration.state)
        else:
            # пропихнуть факультет и статус
            if 'faculty' not in user:
                try:
                    item = db.mail.find_one({'mail': user['st_main']})
                    status = 'employee'
                    if 'status' in user:
                        status = user['status']
                    db.users.update_one(
                        {'chat_id': chat_id},
                        {'$set': {'faculty': item['faculty'], 'status': status}}
                    )
                except Exception as e:
                    db.users.update_one(
                        {'chat_id': chat_id},
                        {'$set': {'faculty': '', 'status': 'employee'}}
                    )
            keyboard = await get_votion_kb(msg)
            await msg.answer(await get_message(msg, 'selection_faculties'), reply_markup=keyboard)
            # await msg.answer('Давай голосовать!')
            await state.set_state(Actions.waiting_for_votion.state)
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def registration_start(msg: types.Message, state: FSMContext):
    try:
        global info
        if info == 0:
            await msg.answer(await get_message(msg, 'voting_end'))
            await state.finish()
            return
        
        if msg.text == await get_button(msg, 'employee'):
            db.users.update_one({'chat_id': msg.chat.id},
                                {'$set': {'status': 'employee'}})
            
            await msg.answer(await get_message(msg, 'st_format'), reply_markup=types.ReplyKeyboardRemove())

        elif msg.text == await get_button(msg, 'foreign_student'):
            db.users.update_one({'chat_id': msg.chat.id},
                                {'$set': {'status': 'foreign_student'}})

            await msg.answer(await get_message(msg, 'registration'), reply_markup=types.ReplyKeyboardRemove())

        else:
            user = db.users.find_one({'chat_id': msg.chat.id})
            data = {}
            if 'status' not in user:
                data['status'] = 'student'

            text = msg.text.lower()
            result = re.match(r'st[0-9]{6,6}', text)
            if result != None and len(text) == 8:
                data['st_main'] = text

                if db.foreign.find_one({'mail': text}):
                    data['status'] = 'foreign_student'

                faculty = ''
                try:
                    faculty = db.mail.find_one({'mail': text})['faculty']
                except Exception as e:
                    if data['status'] != 'foreign_student':
                        data['status'] = 'employee'
                    print(e)
                
                data['faculty'] = faculty

                db.users.update_one({'chat_id': msg.chat.id},
                                    {'$set': data})
                
                await msg.answer(await get_message(msg, 'registration_successfully'), reply_markup=await vote_kb(msg))
                await state.set_state(Actions.waiting_for_votion.state)
            else:
                await msg.answer(await get_message(msg, 'invalid_format'))
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


async def get_votion_kb(msg):
    user = db.users.find_one({'chat_id': msg.from_user.id})

    keyboard = types.InlineKeyboardMarkup()

    faculties = db.faculties.find({})
    buts = []
    ind = -1
    for faculty in faculties:
        if faculty['shot_faculty'] == user['faculty'] or faculty['faculty'] == 'КИО' and user['status'] == 'foreign_student':
            continue
        
        buts.append(types.InlineKeyboardButton(text=NOK + ' ' + faculty['faculty'], callback_data=str(ind)))
        ind -= 1

        if len(buts) == 2:
            keyboard.add(*buts)
            buts = []
    
    if len(buts) != 0:
        keyboard.add(*buts)

    text = await get_button(msg, 'end_voting')
    but = types.InlineKeyboardButton(text=text, callback_data='0')
    keyboard.add(but)

    return keyboard


async def voting_start(msg: types.Message, state: FSMContext):
    try:
        global info
        if info == 0:
            await msg.answer(await get_message(msg, 'voting_end'))
            await state.finish()
            return

        keyboard = await get_votion_kb(msg)
        text = await get_message(msg, 'selection_faculties')
        await msg.answer(text, reply_markup=keyboard)
        await state.set_state(Actions.waiting_for_votion.state)
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


# @dp.callback_query_handler(lambda c:True)
async def inline_callback(call, state: FSMContext):
    try:
        global info
        if info == 0:
            # await call.answer(await get_message(call.message, 'voting_end'))
            await bot.send_message(call.message.chat.id, await get_message(call.message, 'voting_end'))
            await call.message.delete()
            await state.finish()
            return

        keyboard = call['message']['reply_markup']

        data_int = int(call.data)

        n = len(keyboard['inline_keyboard'])

        if data_int == 0:
            chat_id = call.message.chat.id

            faculties = []
            for i in range(n):
                for j in range(len(keyboard['inline_keyboard'][i])):
                    if int(keyboard['inline_keyboard'][i][j]['callback_data']) > 0:
                        faculty = keyboard['inline_keyboard'][i][j]['text'][2:]
                        shot_faculty = db.faculties.find_one({'faculty': faculty})['shot_faculty']
                        faculties.append(shot_faculty)
            
            info = db.info.find_one()
            db.users.update_one(
                {'chat_id': call.message.chat.id},
                {'$set': {'event_' + str(info['active_event']): faculties}}
            )

            await bot.send_message(chat_id, await get_message(call.message, 'voting_successfully'))

            # await call.message.delete()
            await state.finish()

            await call.answer()
            await call.message.delete()
            return

        ok = 0
        for i in range(n):
            for j in range(len(keyboard['inline_keyboard'][i])):
                if int(keyboard['inline_keyboard'][i][j]['callback_data']) > 0:
                    ok += 1
        
        print(ok)
        if ok == 3 and data_int < 0:
            await call.answer(text=await get_message(call.message, 'limit'), show_alert=True)
            return
                
        for i in range(n):
            for j in range(len(keyboard['inline_keyboard'][i])):
                if keyboard['inline_keyboard'][i][j]['callback_data'] == str(data_int):
                    print(data_int)
                    t = keyboard['inline_keyboard'][i][j]['text'][2:]

                    but = 0
                    if data_int < 0:
                        but = types.InlineKeyboardButton(text=OK + ' ' + t, callback_data=str(-data_int))
                    else:
                        but = types.InlineKeyboardButton(text=NOK + ' ' + t, callback_data=str(-data_int))

                    keyboard['inline_keyboard'][i][j] = but

        await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
            message_id=call.message.message_id, reply_markup=keyboard)
    except Exception as e:
        print(e)


async def start_logic(msg: types.Message, state: FSMContext):
    try:
        print(msg)

        info_db = db.info.find_one()
        if not info_db or info_db['active_event'] == 0:
            await msg.answer(await get_message(msg, 'voting_not_running'), reply_markup=await vote_kb(msg))
            return

        user = db.users.find_one({'chat_id': msg.chat.id})

        if user and 'event_' + str(info_db['active_event']) in user:
            await msg.answer(await get_message(msg, 'already_voted'), reply_markup=await vote_kb(msg))
            return

        await msg.answer(await get_message(msg, 'hello'), reply_markup=await vote_kb(msg))

        chat_id = msg.from_user.id
        await state.set_state(Actions.waiting_for_press_start_button.state)
    except Exception as e:
        print(e)
        await msg.answer('Произошла ошибка')


def create_bot_factory():
    storage = MongoStorage(host='localhost', port=27017, db_name='aiogram_fsm')
    dp = Dispatcher(bot, storage=storage)
    # dp = Dispatcher(bot, storage=MemoryStorage())

    dp.register_message_handler(
        show,
        commands='show'
    )

    dp.register_message_handler(
        add_faculty,
        commands='add'
    )

    dp.register_message_handler(
        del_faculty,
        commands='del'
    )

    dp.register_message_handler(
        event_start,
        commands='event_start'
    )

    dp.register_message_handler(
        event_stop,
        commands='event_stop'
    )

    dp.register_message_handler(
        get_result,
        commands='get_result'
    )

    dp.register_message_handler(
        add_correct_mails,
        commands='add_correct'
    )

    dp.register_message_handler(
        push_kio,
        commands='push_kio'
    )

    dp.register_message_handler(
        push_kio_from_db,
        commands='push_kio_from_db'
    )

    dp.register_message_handler(
        check_user,
        commands='check_user',
        state='*'
    )

    dp.register_message_handler(
        del_user,
        commands='del_user'
    )

    dp.register_message_handler(
        del_user_all,
        commands='del_user_all',
        state='*'
    )

    dp.register_message_handler(
        press_start_button,
        state=Actions.waiting_for_press_start_button
    )

    dp.register_message_handler(
        registration_start,
        state=Actions.waiting_for_registration
    )

    dp.register_message_handler(
        voting_start,
        state=Actions.waiting_for_votion
    )

    dp.register_callback_query_handler(
        inline_callback,
        state='*'
    )

    dp.register_message_handler(
        start_logic,
        state='*'
    )

    # dp.register_message_handler(
    #     add_user_to_db,
    #     state='*'
    # )

    start_webhook(
        dispatcher=dp,
        webhook_path=WEBHOOK_PATH,
        on_startup=on_startup,
        on_shutdown=on_shutdown,
        skip_updates=True,
        host=WEBAPP_HOST,
        port=WEBAPP_PORT,
    )


if __name__ == '__main__':
    # asyncio.run(create_bot_factory())
    create_bot_factory()